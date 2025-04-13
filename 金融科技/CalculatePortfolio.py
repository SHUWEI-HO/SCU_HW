import os
import openpyxl
import argparse
from tqdm import tqdm
import statistics
from concurrent.futures import ThreadPoolExecutor
from openpyxl.styles import Alignment


def get_parser():
    """參數解析器"""
    parser = argparse.ArgumentParser("金融科技創新應用")
    parser.add_argument("--input", default="data/")
    parser.add_argument("--element", default="mom")
    parser.add_argument("--sheet-name", default="預測IC")
    return parser.parse_args()


def find_element(worksheet, target):
    """
    在 worksheet 中尋找 target 的位置，回傳 (row, col)；若找不到則回傳 (None, None)
    """
    row_pos, col_pos = None, None
    for row in worksheet.iter_rows(
        min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column
    ):
        for cell in row:
            if str(cell.value) == target:
                row_pos = cell.row
                col_pos = cell.column
                break
        if row_pos is not None and col_pos is not None:
            break
    return row_pos, col_pos


class PortfolioCalculate:
    def __init__(self, input_dir, sheet):
        self.input = input_dir
        self.sheet = sheet
        self.elements = {1: "bm", 2: "size", 3: "mom"}
        self.data_storage = []


    def contrast(self):

        filepath = os.path.join(self.input, "IC.xlsx")
        workbook = openpyxl.load_workbook(filepath, data_only=False)
        worksheet = workbook[self.sheet]

        ref_target = "選絕對值最大"
        ref_row, ref_col = find_element(worksheet, ref_target)
        start_row, start_col = ref_row, ref_col + 1

        start_time_keyword = "2013/12"
        end_time_keyword = "2024/12"
        start_time_row, start_time_col = find_element(worksheet, start_time_keyword)
        end_time_row, end_time_col = find_element(worksheet, end_time_keyword)
        total_data_num = end_time_col - start_time_col + 1

        ws_values = list(worksheet.values)

        data_storage = [
            [
                ws_values[start_time_row + idx][start_time_col - 1 + i]
                for idx in range(3)
            ]
            for i in range(total_data_num)
        ]

        id_storage = [
            [
                abs(ws_values[start_time_row + idx][start_time_col - 1 + i])
                for idx in range(3)
            ]
            for i in range(total_data_num)
        ]

        max_data_num = [row.index(max(row)) + 1 for row in id_storage]
        max_data_value = [
            data_list[num - 1] for num, data_list in zip(max_data_num, data_storage)
        ]

        for idx, (num, value) in tqdm(
            enumerate(zip(max_data_num, max_data_value)),
            total=len(max_data_num),
            desc="Pasting...",
        ):
            cell1 = worksheet.cell(row=start_row, column=start_col + 1 + idx)
            cell1.value = num
            cell1.alignment = Alignment(horizontal="center", vertical="center")

            cell2 = worksheet.cell(row=start_row + 1, column=start_col + 1 + idx)

            cell2.value = self.elements.get(num, "")
            cell2.alignment = Alignment(horizontal="center", vertical="center")

            cell3 = worksheet.cell(row=start_row + 2, column=start_col + 1 + idx)
            cell3.value = value
            cell3.alignment = Alignment(horizontal="center", vertical="center")

        workbook.save(filepath)
        workbook.close()

        return max_data_num, max_data_value

    def load_worksheet(self):

        self.data_dict = {}
        for idx, value in enumerate(self.elements.values()):
            data = openpyxl.load_workbook(os.path.join(self.input, f'{value}.xlsx'),
                                          data_only=True)
            self.data_dict[idx+1] = data

    def process(self, element_id, val, date_str="2013/12"):

        element_name = self.elements.get(element_id, None)
        if not element_name:
            raise ValueError(f"無效的元素編號: {element_id}")

        input_data = self.data_dict[element_id]

        ws = input_data[f"{element_name}補值"]
        next_return = input_data["下個月月報酬補值"]

        _, start_column = find_element(ws, date_str)
        if start_column is None:
            raise ValueError(f"在工作表中找不到日期: {date_str}")

        col_index = start_column - 1

        ws_values = list(ws.values)
        nr_values = list(next_return.values)
        SplitNum = [96, 48, 19, 10, 9, 8, 7, 6, 5, 4, 3, 2, 1]

        def _process_column(col_num, val):
            data = []

            for row_idx in range(1, len(ws_values) - 1):

                key_val = ws_values[row_idx][col_num]

                nr_val = nr_values[row_idx][col_num]
                data.append({key_val: nr_val})

            data.sort(key=lambda d: list(d.keys())[0])

            sorted_nr = [list(item.values())[0] for item in data]
            buy_high, buy_low = [], []
            for sp in SplitNum:
                high_mean = statistics.mean(sorted_nr[-sp:])
                low_mean = statistics.mean(sorted_nr[:sp])
                buy_high.append(high_mean)
                buy_low.append(low_mean)

            results = (
                [low - high for high, low in zip(buy_high, buy_low)]
                if val < 0
                else [high - low for high, low in zip(buy_high, buy_low)]
            )
            return results

        columns_to_process = [col_index]

        with ThreadPoolExecutor(max_workers=os.cpu_count() * 10) as executor:
            results = list(
                executor.map(
                    _process_column, columns_to_process, [val] * len(columns_to_process)
                )
            )
            self.data_storage.extend(results)

        year, month = map(int, date_str.split("/"))
        if month < 12:
            month += 1
        else:
            month = 1
            year += 1
        new_date = f"{year}/{month:02d}"

        return new_date

    def store_data(self):

        filepath = os.path.join(self.input, "IC.xlsx")
        output_wb = openpyxl.load_workbook(filepath)
        worksheet = output_wb[self.sheet]

        ref_target = "投資組合"
        standard_row, standard_col = find_element(worksheet, ref_target)
        if standard_row is None or standard_col is None:
            raise ValueError("找不到 '投資組合' 的位置")

        start_row = standard_row + 2
        start_col = standard_col + 2



        for month_data in tqdm(self.data_storage, total=len(self.data_storage), desc='Storing...'):

            for idx, data in enumerate(month_data):
                worksheet.cell(
                    row=start_row + idx,
                    column=start_col,
                    value=round(float(data), 8),
                )
            start_col += 1
        # self.new_row = start_row
        # self.new_col = start_col + 1

        output_wb.save(filepath)


def main():
    args = get_parser()
    input_dir = args.input
    sheet_name = args.sheet_name
    pc = PortfolioCalculate(input_dir, sheet_name)
    pc.load_worksheet()
    elements_list, value_list = pc.contrast()
    date_str = "2013/12"
    for elem, val in tqdm(zip(elements_list, value_list), total=len(elements_list), desc="Processing"):
        date_str = pc.process(elem, val, date_str)
    
    pc.store_data()

if __name__ == "__main__":
    main()
