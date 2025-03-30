import os
import pandas as pd
from sklearn.linear_model import LinearRegression
import argparse
from tqdm import trange, tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

## 設置基本參數
def get_parser():

    parser = argparse.ArgumentParser('金融科技與創新0325作業')
    parser.add_argument("--input-dir", default="data/", help='資料存放的資料夾位置')
    parser.add_argument("--output", default="data/IC.xlsx", help='最後要填入的EXCEL檔')
    parser.add_argument('--sheet-name', default='OLS_ALL', help='上方EXCEL中指定填入的工作表')

    return parser.parse_args()

# 建立類別
class FinanceLesson0325:

    def __init__(self, output, sheet_name):
        '''
        創建類別主要是讓程式碼看起來較為整潔，
        每個Function都有獨立的功能
        
        '''

        self.output = output
        self.sheet = sheet_name
        

    def store_data(self, results, file_name):

        # 開啟欲存放的EXCEL檔
        data = openpyxl.load_workbook(self.output)

        # 選擇指定工作表 (切記指定的工作必須存在!!!)
        work_sheet = data[self.sheet]

        # 追蹤欲開始填入的位置
        row_pos = None
        col_pos = None

        for row in work_sheet.iter_rows(min_row=1, 
                                        max_row=work_sheet.max_row, 
                                        min_col=1, 
                                        max_col=work_sheet.max_column):
            

            for cell in row:
                if str(cell.value) == file_name:
                    
                    row_pos = cell.row
                    col_pos = cell.column
                    break

        if row_pos == None or col_pos == None:
            raise ValueError(f"{file_name} is not in {self.sheet} worksheet!!!")

        print(f'{file_name}位置: {(row_pos, col_pos)}')

        row_start = row_pos
        col_start = col_pos+1

        # 存入數據
        for idx, element in tqdm(enumerate(results), total=len(results), desc='Pasting...'):

            work_sheet.cell(row=row_start, column=col_start+idx, value=round(float(element),6))

        data.save(self.output)
        data.close()

        print(f'{file_name} IC 新增完成')

    def process(self, inputs, file_name):


        # 讀取 bn補值 (x)
        df_x = pd.read_excel(inputs, sheet_name=f"{file_name}補值")
        df_y = pd.read_excel(inputs, sheet_name=f"{file_name}IC")


        # 轉置(Transform)
        df_x = df_x.T
        df_y = df_y.T

        # 給定 bm補值的行標籤為第1列
        # 給定 bmIC的行標籤為第1列(IC)
        df_x.columns = df_x.iloc[0, :]
        df_y.columns = df_y.iloc[1, :]

        # 留下1999/1 ~ 2025/1的數據(bm) -> 刪除代號、名稱
        # 留下1999/1 ~ 2025/1的數據(IC) -> 刪除Time、IC
        df_x = df_x.iloc[2:, :]
        df_y = df_y.iloc[2:, :]

        results = []

        total_run = len(df_x)-1

        start_time = 166 if file_name == 'mom' else 178

        for idx in trange(start_time, total_run, desc=f'Calculating {file_name} IC'):

            if idx == total_run:
                break

            X = df_x.iloc[:idx, :].values
            Y = df_y.iloc[1:idx+1, :].values

            test_x = df_x.iloc[idx:idx+1, :].values

            reg = LinearRegression()
            reg.fit(X, Y)

            Answer = reg.predict(test_x)
            results.append(Answer.tolist()[0][0])

        self.store_data(results, file_name)


def main():
    args = get_parser()

    output = args.output
    sheet_name = args.sheet_name
    FC = FinanceLesson0325(output=output,
                           sheet_name=sheet_name)

    input_dir = args.input_dir
    file_list = os.listdir(input_dir)

    files_name = ['bm', 'size', 'mom']

    for file in file_list:
        file_name = os.path.splitext(file)[0]
        if file_name in files_name:
            input_path = os.path.join(input_dir, file)
            FC.process(inputs=input_path, file_name=file_name)


if __name__ == '__main__':
    main()
