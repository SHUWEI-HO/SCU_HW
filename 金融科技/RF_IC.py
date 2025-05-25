import os
import argparse
import concurrent.futures
from tqdm import trange, tqdm
import pandas as pd
from sklearn.ensemble import RandomForestRegressor
import openpyxl

def get_parser():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default="data/")
    parser.add_argument("--output", default="data/IC-RF.xlsx")
    parser.add_argument("--sheet-name", default="預測IC")
    return parser.parse_args()

class FinanceLesson0325:
    def __init__(self, output, sheet_name):
        self.output = output
        self.sheet = sheet_name
        self.model = RandomForestRegressor(max_depth=3, random_state=999,
                                           n_estimators=100, n_jobs=-1)

    # --------- 寫檔函式保持不變 ---------
    def store_data(self, results, file_name):
        data = openpyxl.load_workbook(self.output)
        ws = data[self.sheet]

        # 找到 file_name 在哪一格
        row_pos = col_pos = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                if str(cell.value) == file_name:
                    row_pos, col_pos = cell.row, cell.column
                    break
            if row_pos: break
        if row_pos is None:
            raise ValueError(f"{file_name} is not in {self.sheet} worksheet!!!")

        # 依序貼值
        for idx, val in enumerate(tqdm(results, desc=f"Pasting {file_name}"),
                                  start=0):
            ws.cell(row=row_pos, column=col_pos + 1 + idx,
                    value=round(float(val), 6))

        data.save(self.output)
        data.close()

    # --------- 只回傳結果，不寫檔 ---------
    def process(self, inputs, file_name):
        df_x = pd.read_excel(inputs, sheet_name=f"{file_name}補值").T
        df_y = pd.read_excel(inputs, sheet_name=f"{file_name}IC").T
        df_x.columns, df_y.columns = df_x.iloc[0], df_y.iloc[1]
        df_x, df_y = df_x.iloc[2:], df_y.iloc[2:]

        results = []
        total_run = len(df_x) - 1
        start_time = 166 if file_name == "mom" else 178

        for idx in trange(start_time, total_run,
                          desc=f"Calculating {file_name} IC"):
            X   = df_x.iloc[:idx].values
            Y   = df_y.iloc[1:idx+1].values.ravel()
            tx  = df_x.iloc[idx:idx+1].values
            self.model.fit(X, Y)
            results.append(self.model.predict(tx)[0])

        return results   # <── 重要：回傳，不寫檔

# ---------- 包裝函式，將結果傳回 ----------
def process_single_file(args):
    input_path, file_name, output, sheet_name = args
    worker = FinanceLesson0325(output, sheet_name)
    results = worker.process(inputs=input_path, file_name=file_name)
    return file_name, results

def main():
    args = get_parser()
    targets = {"bm", "size", "mom"}
    jobs = [(os.path.join(args.input_dir, f), os.path.splitext(f)[0],
             args.output, args.sheet_name)
            for f in os.listdir(args.input_dir)
            if os.path.splitext(f)[0] in targets]

    # --- 平行計算 ---
    with concurrent.futures.ProcessPoolExecutor() as ex:
        futures = ex.map(process_single_file, jobs)

    # --- 計算完畢，再統一寫檔 ---
    writer = FinanceLesson0325(output=args.output,
                               sheet_name=args.sheet_name)
    for file_name, res in futures:
        writer.store_data(res, file_name)

if __name__ == "__main__":
    main()