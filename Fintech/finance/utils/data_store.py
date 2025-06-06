from openpyxl.worksheet.worksheet import Worksheet
from copy import copy
from typing import List
from tqdm import tqdm
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from typing import Dict, List, Tuple, Optional, Any, List, Sequence

# 將工作表的部分資訊貼到新的工作表，貼在跟舊的工作表一樣的位置
def copy_worksheet_range(
    source_ws: Worksheet,
    target_ws: Worksheet,
    min_row: int = 1,
    max_row: int = 26,
    min_col: int = 1,
    max_col: int = 136,
):

    for col in range(min_col, max_col + 1):
        col_letter = source_ws.cell(row=1, column=col).column_letter
        target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[
            col_letter
        ].width

    for row in range(min_row, max_row + 1):
        if row in source_ws.row_dimensions:
            target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

    for row in source_ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            new_cell = target_ws.cell(
                row=cell.row, column=cell.column, value=cell.value
            )
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    for merged_cell_range in source_ws.merged_cells.ranges:

        minc, minr, maxc, maxr = merged_cell_range.bounds

        if minr >= min_row and maxr <= max_row and minc >= min_col and maxc <= max_col:
            target_ws.merge_cells(
                start_row=minr, start_column=minc, end_row=maxr, end_column=maxc
            )


def _collect_overlapping_merges(
    ws: Worksheet, top: int, left: int, bottom: int, right: int
) -> List[Tuple[int, int, int, int]]:

    overlaps = []
    for rng in ws.merged_cells.ranges:
        minc, minr, maxc, maxr = rng.bounds

        if not (maxr < top or minr > bottom or maxc < left or minc > right):
            overlaps.append((minr, minc, maxr, maxc))
    return overlaps


def copy_to_position(
    src_ws: Worksheet,
    tgt_ws: Worksheet,
    src_min_row: int,
    src_max_row: int,
    src_min_col: int,
    src_max_col: int,
    tgt_start_row: int,
    tgt_start_col: int,
):
    row_offset = tgt_start_row - src_min_row
    col_offset = tgt_start_col - src_min_col


    tgt_top = tgt_start_row
    tgt_left = tgt_start_col
    tgt_bottom = tgt_start_row + (src_max_row - src_min_row)
    tgt_right = tgt_start_col + (src_max_col - src_min_col)

    affected_merges = _collect_overlapping_merges(
        tgt_ws, tgt_top, tgt_left, tgt_bottom, tgt_right
    )
    for minr, minc, maxr, maxc in affected_merges:
        tgt_ws.unmerge_cells(
            start_row=minr, start_column=minc, end_row=maxr, end_column=maxc
        )


    for c in range(src_min_col, src_max_col + 1):
        src_letter = src_ws.cell(row=1, column=c).column_letter
        tgt_letter = tgt_ws.cell(row=1, column=c + col_offset).column_letter
        tgt_ws.column_dimensions[tgt_letter].width = src_ws.column_dimensions[
            src_letter
        ].width


    for r in range(src_min_row, src_max_row + 1):
        if r in src_ws.row_dimensions:
            tgt_ws.row_dimensions[r + row_offset].height = src_ws.row_dimensions[
                r
            ].height


    for row in src_ws.iter_rows(
        min_row=src_min_row,
        max_row=src_max_row,
        min_col=src_min_col,
        max_col=src_max_col,
    ):
        for cell in row:
            new_r = cell.row + row_offset
            new_c = cell.column + col_offset
            tgt_cell = tgt_ws.cell(row=new_r, column=new_c)


            tgt_cell.value = cell.value

            if cell.has_style:
                tgt_cell.font = copy(cell.font)
                tgt_cell.border = copy(cell.border)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.number_format = copy(cell.number_format)
                tgt_cell.protection = copy(cell.protection)
                tgt_cell.alignment = copy(cell.alignment)


    for rng in src_ws.merged_cells.ranges:
        minc, minr, maxc, maxr = rng.bounds

        if (
            src_min_row <= minr <= src_max_row
            and src_min_col <= minc <= src_max_col
            and src_min_row <= maxr <= src_max_row
            and src_min_col <= maxc <= src_max_col
        ):
            tgt_ws.merge_cells(
                start_row=minr + row_offset,
                start_column=minc + col_offset,
                end_row=maxr + row_offset,
                end_column=maxc + col_offset,
            )


def storage(
    ws: Worksheet,
    results: List[List],
    work_name: str,
    row_s: int,
    col_s: int,
    shape_vertical: bool = False,
):
    if not shape_vertical:
        # for m_idx, elements in tqdm(
        #     enumerate(results), total=len(results), desc=f"{work_name}"
        # ):
        for m_idx, elements in enumerate(results):
            for idx, element in enumerate(elements):
                element = (
                    round(float(element), 8)
                    if isinstance(element, float)
                    else element
                )
                cell = ws.cell(row=row_s + idx, column=col_s + m_idx)
                cell.value = element
                cell.font = Font(name="Calibri", size=12, bold=True)
    else:
        # for m_idx, elements in tqdm(
        #     enumerate(results), total=len(results), desc=f"{work_name}"
        # ):
        for m_idx, elements in enumerate(results):
            for idx, element in enumerate(elements):
                element = (
                    round(float(element), 8)
                    if isinstance(element, float)
                    else element
                )
                cell = ws.cell(row=row_s + m_idx, column=col_s + idx)
                cell.value = element
                cell.font = Font(name="Calibri", size=12, bold=True)


def modify_block(ws: Worksheet) -> None:
    user_font = Font(
        name="Calibri",
        size=12,
        bold=True,
        italic=False,
        color="000000",
    )

    thick = Side(style="thick")
    thick_border = Border(top=thick, bottom=thick, left=thick, right=thick)
    center_align = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(
        min_row=ws.min_row,
        max_row=ws.max_row,
        min_col=ws.min_column,
        max_col=ws.max_column,
    ):
        for cell in row:

            if cell.value not in (None, ""): 
                cell.border = thick_border
                new_font = copy(cell.font)
                for attr in (
                    "name",
                    "size",
                    "bold",
                    "italic",
                    "color",
                    "underline",
                    "strike",
                    "vertAlign",
                    "charset",
                    "scheme",
                ):
                    val = getattr(user_font, attr)
                    if val is not None:
                        setattr(new_font, attr, val)
                new_font.bold = True
                cell.font = new_font

                cell.alignment = center_align


def type_block(ws: Worksheet, value: Any, row: int, col: int) -> None:

    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = Font(name="Calibri", size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def fill_color(
    ws: Worksheet,
    row_s: int,
    col_s: int,
    row_e: int,
    col_e: int,
    color: str = "FFFFE699",
) -> None:

    for r in range(row_s, row_e+1):
        for c in range(col_s, col_e+1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(fill_type="solid", fgColor=color)
