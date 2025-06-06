import os
import itertools
import statistics
import numpy as np
from typing import Dict, List, Tuple, Optional, Any, Sequence
from concurrent.futures import ThreadPoolExecutor
from statistics import mean, stdev
from pathlib import Path
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm, trange
from utils.data_store import copy_worksheet_range, copy_to_position, storage, modify_block, type_block, fill_color


ELEMENT_ID: Dict[int, str] = {1: "bm", 2: "size", 3: "mom"}
SPLIT_NUM: List[int] = [96, 48, 19, 10, 9, 8, 7, 6, 5, 4, 3, 2, 1]
MODEL_NAME: Dict[str, str] = {"LR": "OLS", "RF": "RF", "NN": "NN"}


def find_element(
    ws: Worksheet, 
    target: str
) -> Tuple[Optional[int], Optional[int]]:
    
    row_pos = None
    col_pos = None
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_values = [str(v) for v in row]
        if target in row_values:
            row_pos = r_idx
            col_pos = row_values.index(target) + 1
            break
    return row_pos, col_pos


def select_factor(data: Dict[str, List[float]]) -> Tuple[List[List[float]], List[int], List[float]]:
    
    lengths = {len(data[k]) for k in ELEMENT_ID.values()}
    if len(lengths) != 1:
        raise ValueError("計算錯誤，請檢查預測是否有誤")

    data_storage = [
        [bm, size, mom] for bm, size, mom in zip(data["bm"], data["size"], data["mom"])
    ]

    id_storage = []
    select_value = []
    for rec in tqdm(data_storage, desc="篩選因子"):
        idx: int = max(range(3), key=lambda i: abs(rec[i])) + 1
        id_storage.append(idx)
        select_value.append(float(rec[idx - 1]))

    return data_storage, id_storage, select_value


def load_worksheet(inputs: str, output: str = None, element_id: List = None) -> Dict[str, Workbook]:
    
    if element_id is None:
        if isinstance(ELEMENT_ID, dict):
            data_dict: Dict[str, Workbook] = {
                v: openpyxl.load_workbook(os.path.join(inputs, f"{v}.xlsx"), data_only=True)
                for v in ELEMENT_ID.values()
            }

            data_dict["IC"] = openpyxl.load_workbook(
                output, data_only=True
            )
    else:
        if isinstance(element_id, list):
            data_dict: Dict[str, Workbook] = {
                v: openpyxl.load_workbook(os.path.join(inputs, f"{v}.xlsx"), data_only=True)
                for v in element_id
            }


    return data_dict


def portfolio_calculate(
    data_dict: Dict[str, Workbook],
    value: float,
    element_id: int,
    data_storage: List[List[float]],
    date_str: str = "2013/12",
    ) -> Tuple[str, List[List[float]]]:
    
    element_name = ELEMENT_ID[element_id]
    wb = data_dict[element_name]
    ws = wb[f"{element_name}補值"]
    nr = wb["下個月月報酬補值"]

    _, col = find_element(ws, date_str)
    if col is None:
        raise ValueError(f"在工作表中找不到日期: {date_str}")
    col -= 1

    ws_values = list(ws.values)
    nr_values = list(nr.values)

    def _process(col_idx: int, v: float) -> List[float]:
        pairs = sorted(
            (
                (ws_values[r][col_idx], nr_values[r][col_idx])
                for r in range(1, len(ws_values) - 1)
            ),
            key=lambda p: p[0],
        )
        sorted_nr = [p[1] for p in pairs]
        high_mean = [statistics.mean(sorted_nr[-sp:]) for sp in SPLIT_NUM]
        low_mean = [statistics.mean(sorted_nr[:sp]) for sp in SPLIT_NUM]
        return [(h - l) if v >= 0 else (l - h) for h, l in zip(high_mean, low_mean)]

    with ThreadPoolExecutor(max_workers=os.cpu_count() * 100) as ex:
        data_storage.extend(ex.map(_process, [col], [value]))

    y, m = map(int, date_str.split("/"))
    m = m + 1 if m < 12 else 1
    y = y if m != 1 else y + 1
    next_date: str = f"{y}/{m:02d}"
    return next_date, data_storage


def win_rate(
    portfolio_data: List[List[float]],
    market: List[float],
    ) -> Tuple[List[List[int]], List[int], List[float]]:
    
    if len(portfolio_data) != len(market):
        raise ValueError("數據不對稱，請檢查數據是否筆數一致")

    pf_vs_mk = [
        [int(p > mk) for p in pf]
        for pf, mk in tqdm(
            zip(portfolio_data, market),
            total=len(market),
            desc="計算勝率",
        )
    ]

    totals = [sum(col) for col in zip(*pf_vs_mk)]
    rates = [t / len(pf_vs_mk) for t in totals]
    return pf_vs_mk, totals, rates


def cumulative_reward(data: List[List[float]]) -> List[List[float]]:
    col_iter = zip(*data)
    return [
        list(itertools.accumulate(col))
        for col in tqdm(
            col_iter, total=len(data[0]), desc="計算累積報酬"
        )
    ]


def sharpe_ratio(data: Sequence[Sequence[float]], market: List[float]) -> List[float]:

    sharpe = []

    for col in zip(*data):

        means = mean(col)
        standard = stdev(col)

        sharpe.append([means, standard, means / standard if standard else float("nan")])  

    market_means = mean(market)
    market_standard = stdev(market)
    market_sharpe_ratio = (
        market_means / market_standard if market_standard else float("nan")
    )
    
    sharpe_market = []
    sharpe_market.append([market_means, market_standard, market_sharpe_ratio])

    return sharpe, sharpe_market


def ensure_sheet(wb: Workbook, sheet_name: str) -> Worksheet:

    if sheet_name not in wb.sheetnames:
        return wb.create_sheet(sheet_name)
    return wb[sheet_name]

def r_square(real: List[List[float]], pred: List[List[float]]) -> List[List[float]]:

    real = np.array(real)
    pred = np.array(pred)

    rs_rp = np.square(pred - real)
    rs_rr = np.square(real)

    RS = 1 - float(np.sum(rs_rp)/ np.sum(rs_rr))

    return rs_rp.tolist() , rs_rr.tolist(), RS

def classifier_store(
    predict_cumulative_reward: List[List[List]], 
    real_cumulative_reward: List[List],
    market: List,
    element_id: List,
    output: Path,
    date: List
)-> None:
    

    # 創建新的工作表
    wb = Workbook()
    default_sheet = wb.active 
    wb.remove(default_sheet)

    select_dict = {0:10, 1:20, 2:50, 3:100, 8:192, 12:964}

    # 不需要將13次的分類，只需要取10、20、50、100、192跟964的等分的篩選
    for idx, (pcrs, rcr) in tqdm(enumerate(zip(zip(*predict_cumulative_reward), 
                             real_cumulative_reward)),
                             total=len(real_cumulative_reward),
                             desc="將所有預測即真實累積報酬分類"):
        
        if idx in select_dict.keys():
            
            ws = wb.create_sheet(title=f'{select_dict[idx]}等分畫圖')

            ws.merge_cells('A1:B1')
            type_block(ws, '投資組合', 1, 1)
            type_block(ws, '公司數', 1, 3)
            type_block(ws, '964', 1, 4)
            type_block(ws, f'{select_dict[idx]}等分', 2, 1)
            type_block(ws, '時間', 2, 2)


            storage(ws, [date], '填入時間', 2, 3, shape_vertical=True)
            storage(ws, [element_id], '填入模型名稱', 3, 2)

            type_block(ws, 'Real', 10, 2)
            type_block(ws, '大盤', 11, 2)

            for p_idx, pcr in enumerate(pcrs):
                storage(ws, [pcr], f'填入{element_id[p_idx]}的{select_dict[idx]}等分', 3+p_idx, 3, shape_vertical=True)
                fill_color(ws, 3, 3, 9, 136, color="FFFDE9D9")


            storage(ws, [rcr], f"填入真實累積報酬的{select_dict[idx]}等分", 10, 3, shape_vertical=True)
            fill_color(ws, 10, 3, 10, 136, color="FFDAEEF3")
            storage(ws, [market], f"填入大盤累積報酬", 11, 3, shape_vertical=True)
            fill_color(ws, 11, 3, 11, 136, color="FFDAEEF3")

            modify_block(ws)

        
    wb.save(output)
    

        


def data_store( file: Path, data: dict[str, Any], model: str) -> None:

    excel_data = openpyxl.load_workbook(file)

    ws_gt = ensure_sheet(excel_data, "真實IC")
    ws_ic = ensure_sheet(excel_data, "預測IC")
    ws_wr = ensure_sheet(excel_data, f"{MODEL_NAME[model]}勝率")
    ws_cr = ensure_sheet(excel_data, f"{MODEL_NAME[model]}累積報酬")
    ws_sp = ensure_sheet(excel_data, f"{MODEL_NAME[model]}夏普比率")
    ws_r2 = ensure_sheet(excel_data, f"{MODEL_NAME[model]}樣本外R2")

    IC = data["IC"]                                     # 三因子IC數據
    ID = data["ID"]                                     # 每個數據最後選出來的因子
    SV = data["select_values"]                          # 篩選出的因子的數據
    PD = data["portfolio_data"]                         # 所有投資組合數據
    PV = data["portfolio_value"]                        # 所有投資組合數據分類
    TC = data["total_count"]                            # 所有勝過大盤的數據總數
    WR = data["win_rate"]                               # 所有投資組合的勝率
    CR = data["cumulative_rewards"]                     # 所有投資組合的累積報酬率
    SP = data["sharpe_ratio"]                           # 夏普比率數據
    SM = data["sharpe_ratio_market"]                    # 夏普比率數據(大盤)
    PR = data["PR_RS"]                                  # (預測-真實)^2
    RR = data["RR_RS"]                                  # (真實)^2)
    RS = data["RS"]                                     # 整體樣本外R^2

    # 整合篩選因子的數據並儲存
    factors = [[i, ELEMENT_ID[i], sv] for i, sv in zip(ID, SV)]

    # 儲存基本IC資訊
    storage(ws_ic, IC, "儲存IC至工作表->預測IC", 3, 3)
    storage(ws_ic, factors, "儲存篩選因子至工作表->預測IC", 6, 3)
    storage(ws_ic, PD, "儲存投資組合至工作表->預測IC", 12, 3)

    # 將以填入過的數據複製到其他工作表中，這樣就不需要在重複填上數據，不然會很亂
    copy_worksheet_range(ws_ic, ws_wr, 1, 26, 1, 136)
    copy_worksheet_range(ws_ic, ws_cr, 1, 26, 1, 136)

    # 儲存勝率相關資訊
    storage(ws_wr, PV, f"將投資組合評斷分數儲存至工作表->{model}勝率", 12, 3)
    storage(ws_wr, [TC], f"將各等分數勝過大盤總數儲存至工作表->{model}勝率", 12, 137)
    storage(ws_wr, [WR], f"將各等分勝率儲存至工作表->{model}勝率", 12, 138)
    type_block(ws_wr, "sum", 11, 137)
    type_block(ws_wr, "勝率", 11, 138)

    # 儲存累積報酬相關資訊，因為這個與前面的形狀不同所以貼到工作表的位置需要做更正
    storage(ws_cr, CR, f"將各等分數累積報酬儲存至工作表->{model}累積報酬", 12, 3, shape_vertical=True)

    """
    儲存夏普比率
    先複製預測IC的工作表資訊，在建立欄位名稱: 平均值、標準差以及夏普比率
    """
    copy_worksheet_range(ws_ic, ws_sp, 1, 26, 1, 136)
    type_block(ws_sp, "平均值", 11, 137)
    type_block(ws_sp, "標準差", 11, 138)
    type_block(ws_sp, "夏普比率", 11, 139)
    storage(ws_sp, SP, f"將夏普比率儲存至工作表->{model}夏普比率", 12, 137, shape_vertical=True)
    storage(ws_sp, SM, f"將大盤的夏普比率儲存至工作表->{model}夏普比率", 26, 137, shape_vertical=True)

    # 儲存R-square
    copy_to_position(ws_gt, ws_r2, 1, 5, 1, 136, 7, 1)              # 複製真實IC
    copy_to_position(ws_ic, ws_r2, 1, 5, 1, 136, 1, 1)              # 複製預測IC
    copy_to_position(ws_ic, ws_r2, 3, 5, 2, 2, 13, 2)
    copy_to_position(ws_ic, ws_r2, 3, 5, 2, 2, 17, 2)
    storage(ws_r2, PR, f"將(預測-真實)^2 R-square儲存至工作表->{model}樣本外R2", 13, 3)
    storage(ws_r2, RR, f"將(真實^2) R-square儲存至工作表->{model}樣本外R2", 17, 3)
    fill_color(ws_r2, 13, 3, 15, 136, color="FFFFE699")
    fill_color(ws_r2, 17, 3, 19, 136, color="FFFFE699")
    type_block(ws_r2, "(預測-真實)^2", 13, 1)
    type_block(ws_r2, "真實^2", 17, 1)
    type_block(ws_r2, "R^2", 21, 1)
    type_block(ws_r2, RS, 21, 2)

    # 將所有的工作表有數據的部分都能擁有粗框線
    for ws in [ws_ic, ws_cr, ws_wr, ws_sp, ws_r2]:
        modify_block(ws)

    excel_data.save(file)
    excel_data.close()

    print("儲存成功!")
